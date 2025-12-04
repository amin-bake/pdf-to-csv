"""
File Converters Module
Handles conversion of extracted table data to various output formats (CSV, Excel, JSON, Text).
"""
import os
import csv
import json
import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from analyzers import (
    analyze_table_structure,
    create_headers,
    validate_table_data
)
from extractors import extract_structured_text_json


def save_tables_to_text(tables, output_dir, base_filename, merge=False, pdf_path=None):
    """
    Save extracted content to plain text (.txt) files.
    For documents with tables, extracts table data.
    For text documents (CVs, resumes), extracts full text content.
    Returns list of created file paths.
    """
    converted_files = []
    
    # Check if this is valid tabular data or just text
    is_valid_table_data = validate_table_data(tables)
    
    if not is_valid_table_data and pdf_path:
        # Extract as plain text for non-tabular documents
        output_path = os.path.join(output_dir, f"{base_filename}.txt")
        
        with pdfplumber.open(pdf_path) as pdf:
            full_text = []
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ''
                if text.strip():
                    full_text.append(f"=== Page {page_num} ===")
                    full_text.append(text)
                    full_text.append("")  # Empty line between pages
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(full_text))
        
        converted_files.append(output_path)
    elif tables:
        # Extract tables as formatted text
        if merge:
            # Merge all tables into a single text file
            output_path = os.path.join(output_dir, f"{base_filename}.txt")
            with open(output_path, 'w', encoding='utf-8') as f:
                for table_idx, table in enumerate(tables, start=1):
                    if len(tables) > 1:
                        f.write(f"=== Table {table_idx} ===\n\n")
                    
                    # Calculate column widths for alignment
                    if table:
                        col_widths = [0] * max(len(row) for row in table)
                        for row in table:
                            for col_idx, cell in enumerate(row):
                                col_widths[col_idx] = max(col_widths[col_idx], len(str(cell)))
                        
                        # Write rows with aligned columns
                        for row in table:
                            row_text = []
                            for col_idx, cell in enumerate(row):
                                cell_text = str(cell).ljust(col_widths[col_idx])
                                row_text.append(cell_text)
                            f.write('  '.join(row_text) + '\n')
                        
                        if table_idx < len(tables):
                            f.write('\n')  # Separator between tables
            
            converted_files.append(output_path)
        else:
            # Save each table as a separate text file
            for idx, table in enumerate(tables, start=1):
                output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.txt")
                with open(output_path, 'w', encoding='utf-8') as f:
                    if table:
                        # Calculate column widths for alignment
                        col_widths = [0] * max(len(row) for row in table)
                        for row in table:
                            for col_idx, cell in enumerate(row):
                                col_widths[col_idx] = max(col_widths[col_idx], len(str(cell)))
                        
                        # Write rows with aligned columns
                        for row in table:
                            row_text = []
                            for col_idx, cell in enumerate(row):
                                cell_text = str(cell).ljust(col_widths[col_idx])
                                row_text.append(cell_text)
                            f.write('  '.join(row_text) + '\n')
                
                converted_files.append(output_path)
    
    return converted_files


def save_tables_to_csv(tables, output_dir, base_filename, merge=False):
    """
    Save extracted tables to CSV files.
    Returns list of created file paths.
    """
    converted_files = []
    
    if merge:
        # Merge all tables into a single CSV
        output_path = os.path.join(output_dir, f"{base_filename}.csv")
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for table in tables:
                for row in table:
                    writer.writerow(row)
        converted_files.append(output_path)
    else:
        # Save each table as a separate CSV
        for idx, table in enumerate(tables, start=1):
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.csv")
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in table:
                    writer.writerow(row)
            converted_files.append(output_path)
    
    return converted_files


def save_tables_to_excel(tables, output_dir, base_filename, merge=False):
    """
    Save extracted tables to Excel (.xlsx) files.
    Returns list of created file paths.
    """
    converted_files = []
    
    if merge:
        # Merge all tables into a single Excel file with one sheet containing all rows
        output_path = os.path.join(output_dir, f"{base_filename}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Data"
        
        # Append all rows from all tables sequentially
        for table in tables:
            for row in table:
                ws.append(row)
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        converted_files.append(output_path)
    else:
        # Save each table as a separate Excel file
        for idx, table in enumerate(tables, start=1):
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Table Data"
            
            # Append all rows from the table
            for row in table:
                ws.append(row)
            
            # Auto-adjust column widths
            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            converted_files.append(output_path)
    
    return converted_files


def save_tables_to_json(tables, output_dir, base_filename, merge=False, pdf_path=None):
    """
    Save extracted tables to JSON files with intelligent structure detection.
    Falls back to structured text extraction for non-tabular documents.
    Returns list of created file paths.
    """
    converted_files = []
    
    if merge:
        # Merge all tables into a single JSON file with table metadata
        output_path = os.path.join(output_dir, f"{base_filename}.json")
        
        # Check if extracted data is truly tabular
        is_valid_table_data = validate_table_data(tables)
        
        # If tables look like poorly parsed text, fall back to text extraction
        if not is_valid_table_data:
            if pdf_path and os.path.exists(pdf_path):
                result = extract_structured_text_json(pdf_path)
            else:
                result = {"tables": []}
            
            # Write the text extraction result and skip table processing
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            converted_files.append(output_path)
        else:
            # Strategy: Find the first table with valid headers, use those headers for ALL subsequent tables
            master_headers = None
            all_data_rows = []
            
            for table_idx, table in enumerate(tables, start=1):
                if len(table) < 1:
                    continue
                
                # Analyze table structure
                structure = analyze_table_structure(table)
                
                if not structure:
                    continue
                
                # If we don't have master headers yet, try to get them from this table
                if master_headers is None and structure['header_row_idx'] is not None:
                    header_row_idx = structure['header_row_idx']
                    master_headers = create_headers(table[header_row_idx], structure['column_count'], structure)
                    
                    # Get data from this table (if any)
                    data_start_idx = structure['data_start_idx']
                    if data_start_idx < len(table):
                        for row_idx in range(data_start_idx, len(table)):
                            row = table[row_idx]
                            if any(str(cell).strip() for cell in row if cell):
                                all_data_rows.append(row)
                else:
                    # We already have master headers, just collect data rows
                    # Start from the beginning if no headers found in this table
                    if structure['header_row_idx'] is None:
                        start_idx = 0
                    else:
                        # Skip the header row in this table since we're using master headers
                        start_idx = structure['data_start_idx'] if structure['data_start_idx'] < len(table) else 0
                    
                    for row_idx in range(start_idx, len(table)):
                        row = table[row_idx]
                        # Skip empty rows
                        if not any(str(cell).strip() for cell in row if cell):
                            continue
                        # Skip rows that look like headers (contain header keywords)
                        row_str = ' '.join([str(cell).lower() for cell in row if cell])
                        if any(keyword in row_str for keyword in ['sno', 'barcode', 'product', 'image', 'brand', 'description', 'wholesale', 'retail']):
                            # Check if it's actually a header row (not data with these words in product names)
                            is_header = False
                            for cell in row[:3]:  # Check first 3 cells
                                cell_str = str(cell).lower().strip()
                                if cell_str in ['sno', 's.no', 'no', 'barcode', 'bar code', 'product image', 'image']:
                                    is_header = True
                                    break
                            if is_header:
                                continue
                        
                        all_data_rows.append(row)
            
            # If we found headers and data, create the merged result
            has_valid_tables = False
            if master_headers and all_data_rows:
                has_valid_tables = True
                
                # Convert all data rows to dictionaries using master headers
                table_data = []
                for row in all_data_rows:
                    row_dict = {}
                    for col_idx, header in enumerate(master_headers):
                        value = row[col_idx] if col_idx < len(row) else ""
                        value = str(value).strip() if value else ""
                        row_dict[header] = value
                    
                    table_data.append(row_dict)
                
                result = {
                    "tables": [{
                        "table_number": 1,
                        "rows": len(table_data),
                        "columns": len(master_headers),
                        "headers": master_headers,
                        "data": table_data
                    }]
                }
            else:
                result = {"tables": []}
            
            # If no valid tables found in table mode, extract as structured text
            if not has_valid_tables or not result["tables"]:
                if pdf_path and os.path.exists(pdf_path):
                    result = extract_structured_text_json(pdf_path)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            converted_files.append(output_path)
    else:
        # Save each table as a separate JSON file
        has_valid_tables = False
        
        for idx, table in enumerate(tables, start=1):
            if len(table) < 2:
                continue
            
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.json")
            
            # Analyze table structure intelligently
            structure = analyze_table_structure(table)
            
            if not structure or structure['header_row_idx'] is None:
                continue
            
            has_valid_tables = True
            header_row_idx = structure['header_row_idx']
            data_start_idx = structure['data_start_idx']
            
            if data_start_idx >= len(table):
                continue
            
            # Extract title if present
            title_text = None
            if structure['title_rows']:
                title_parts = []
                for title_analysis in structure['title_rows']:
                    non_empty = [cell for cell in title_analysis['cells'] if cell]
                    title_parts.extend(non_empty)
                title_text = ' '.join(title_parts) if title_parts else None
            
            # Create headers
            headers = create_headers(table[header_row_idx], structure['column_count'], structure)
            
            # Convert data rows to dictionaries
            table_data = []
            for row_idx in range(data_start_idx, len(table)):
                row = table[row_idx]
                
                # Skip empty rows
                if not any(str(cell).strip() for cell in row if cell):
                    continue
                
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    value = row[col_idx] if col_idx < len(row) else ""
                    value = str(value).strip() if value else ""
                    row_dict[header] = value
                
                table_data.append(row_dict)
            
            if table_data:
                result = {
                    "table_number": idx,
                    "rows": len(table_data),
                    "columns": len(headers),
                    "headers": headers,
                    "data": table_data
                }
                
                if title_text:
                    result["title"] = title_text
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(result, f, indent=2, ensure_ascii=False)
                
                converted_files.append(output_path)
        
        # If no valid tables found, create a single text file
        if not has_valid_tables and pdf_path and os.path.exists(pdf_path):
            output_path = os.path.join(output_dir, f"{base_filename}.json")
            result = extract_structured_text_json(pdf_path)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            converted_files.append(output_path)
    
    return converted_files
