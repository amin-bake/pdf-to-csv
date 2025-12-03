"""
Conversion Service
Handles PDF to CSV and Excel conversion using pdfplumber and tabula.
Port: 5002
"""
import os
import uuid
import csv
import json
import threading
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
import pdfplumber
import tempfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# Configuration
UPLOAD_FOLDER = os.path.join(tempfile.gettempdir(), 'pdf-to-csv-uploads')
CONVERTED_FOLDER = os.path.join(tempfile.gettempdir(), 'pdf-to-csv-converted')

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CONVERTED_FOLDER, exist_ok=True)

# In-memory storage for conversion jobs
conversion_jobs = {}


def extract_tables_pdfplumber(pdf_path):
    """
    Extract tables from PDF using pdfplumber.
    Returns list of tables (each table is a list of rows).
    """
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_tables = page.extract_tables()
            if page_tables:
                for table in page_tables:
                    # Clean table: replace None with empty string
                    clean_table = [
                        [cell if cell is not None else "" for cell in row]
                        for row in table
                    ]
                    tables.append(clean_table)
    return tables


def extract_text_lines(pdf_path):
    """
    Fallback: Extract structured text when no tables found.
    Returns structured data for non-tabular documents (CVs, reports, etc.).
    """
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ''
            for line in text.splitlines():
                if line.strip():
                    lines.append([line])
    return [lines] if lines else []


def extract_structured_text_json(pdf_path):
    """
    Extract structured text content for JSON output (CVs, resumes, reports).
    Organizes content by pages and sections.
    """
    result = {
        "document_type": "text",
        "pages": []
    }
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ''
            
            if not text.strip():
                continue
            
            # Split into lines and organize
            lines = [line for line in text.splitlines() if line.strip()]
            
            page_data = {
                "page_number": page_num,
                "line_count": len(lines),
                "content": text.strip(),
                "lines": lines
            }
            
            result["pages"].append(page_data)
    
    return result


def save_tables_to_csv(tables, output_dir, base_filename, merge=False):
    """
    Save extracted tables to CSV files.
    Returns list of created file paths.
    """
    converted_files = []
    
    if merge:
        # Merge all tables into a single CSV
        output_path = os.path.join(output_dir, f"{base_filename}.csv")
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for table in tables:
                for row in table:
                    writer.writerow(row)
        converted_files.append(output_path)
    else:
        # Save each table as a separate CSV
        for idx, table in enumerate(tables, start=1):
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.csv")
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in table:
                    writer.writerow(row)
            converted_files.append(output_path)
    
    return converted_files


def save_tables_to_excel(tables, output_dir, base_filename, merge=False):
    """
    Save extracted tables to Excel (.xlsx) files.
    Returns list of created file paths.
    """
    converted_files = []
    
    if merge:
        # Merge all tables into a single Excel file with one sheet containing all rows
        output_path = os.path.join(output_dir, f"{base_filename}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Data"
        
        # Append all rows from all tables sequentially
        for table in tables:
            for row in table:
                ws.append(row)
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        converted_files.append(output_path)
    else:
        # Save each table as a separate Excel file
        for idx, table in enumerate(tables, start=1):
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Table Data"
            
            # Append all rows from the table
            for row in table:
                ws.append(row)
            
            # Auto-adjust column widths
            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            converted_files.append(output_path)
    
    return converted_files


def save_tables_to_json(tables, output_dir, base_filename, merge=False, pdf_path=None):
    """
    Save extracted tables to JSON files with intelligent structure detection.
    Falls back to structured text extraction for non-tabular documents.
    Returns list of created file paths.
    """
    
    def clean_header(header):
        """Clean and normalize header text."""
        if not header or not str(header).strip():
            return None
        # Remove newlines and extra whitespace
        cleaned = str(header).replace('\n', ' ').strip()
        # Remove multiple spaces
        cleaned = ' '.join(cleaned.split())
        return cleaned if cleaned else None
    
    def is_title_row(row):
        """
        Detect if a row is likely a title/header row rather than column headers.
        Title rows typically have:
        1. Only one or two cells with content (rest are empty)
        2. Longer text that spans conceptually across the table
        3. No data-like patterns (numbers, codes, structured text)
        """
        non_empty_cells = [cell for cell in row if cell and str(cell).strip()]
        
        # If most cells are empty, likely a title
        if len(non_empty_cells) <= 2 and len(row) > 3:
            # Check if the non-empty content is longer text (likely a title)
            for cell in non_empty_cells:
                cell_text = str(cell).strip()
                # Titles are usually longer and contain spaces or newlines
                if len(cell_text) > 15 or '\n' in str(cell):
                    return True
        
        return False
    
    def analyze_table_structure(table):
        """
        Intelligently analyze table structure to identify:
        - Title rows
        - Header rows
        - Data rows
        - Empty/separator rows
        Returns structured information about the table.
        """
        if not table or len(table) < 2:
            return None
        
        structure = {
            'title_rows': [],
            'header_row_idx': None,
            'data_start_idx': None,
            'column_count': 0,
            'has_sequential_ids': False
        }
        
        # Analyze each row
        row_analysis = []
        for idx, row in enumerate(table):
            analysis = {
                'idx': idx,
                'type': 'unknown',
                'non_empty_count': 0,
                'numeric_count': 0,
                'text_count': 0,
                'long_text_count': 0,
                'barcode_count': 0,
                'short_number_count': 0,
                'is_empty': False,
                'cells': []
            }
            
            for cell in row:
                cell_str = str(cell).strip() if cell else ""
                analysis['cells'].append(cell_str)
                
                if not cell_str:
                    continue
                
                analysis['non_empty_count'] += 1
                
                # Check cell characteristics
                if cell_str.isdigit():
                    analysis['numeric_count'] += 1
                    # Long numbers might be barcodes (data)
                    if len(cell_str) > 8:
                        analysis['barcode_count'] += 1
                    # Short numbers (1-3 digits) might be IDs or quantities (data)
                    elif len(cell_str) <= 3:
                        analysis['short_number_count'] += 1
                elif any(c.isalpha() for c in cell_str):
                    analysis['text_count'] += 1
                    if len(cell_str) > 20 or '\n' in cell_str:
                        analysis['long_text_count'] += 1
            
            # Classify row type
            if analysis['non_empty_count'] == 0:
                analysis['type'] = 'empty'
                analysis['is_empty'] = True
            elif analysis['non_empty_count'] <= 2 and len(row) > 3 and analysis['long_text_count'] > 0:
                analysis['type'] = 'title'
            else:
                # Determine if it's a header or data row
                analysis['type'] = 'unknown'
            
            row_analysis.append(analysis)
        
        # Find title rows (at the beginning)
        for analysis in row_analysis:
            if analysis['type'] == 'title':
                structure['title_rows'].append(analysis)
            elif analysis['type'] != 'empty':
                break  # Stop at first non-title, non-empty row
        
        # Find header row using keyword scoring
        header_candidates = []
        header_keywords = [
            'sno', 's.no', 'no', 'serial', 'number', '#', 'name', 'description', 
            'price', 'code', 'barcode', 'bar code', 'brand', 'item', 
            'product', 'quantity', 'qty', 'amount', 'date', 'time', 
            'category', 'type', 'status', 'id', 'image', 'wholesale', 
            'retail', 'ml', 'pcs', 'ctn', 'carton', 'pieces', 'count',
            'total', 'subtotal', 'unit', 'size', 'color', 'model', 'sku'
        ]
        
        for analysis in row_analysis:
            if analysis['type'] in ['title', 'empty']:
                continue
            
            score = 0
            is_likely_data = False
            
            # Check each cell for header vs data characteristics
            for cell in analysis['cells'][:8]:  # Check first 8 columns
                if not cell:
                    continue
                
                cell_lower = cell.lower().strip()
                
                # Strong indicators this is a HEADER row
                if any(keyword == cell_lower or keyword in cell_lower for keyword in header_keywords):
                    score += 5  # Strong header signal
                
                # Check for typical header patterns
                if any(c.isalpha() for c in cell) and len(cell) >= 2 and len(cell) <= 30:
                    # Contains letters, reasonable length for a header
                    if not cell.isdigit() and not (len(cell) > 8 and cell.replace('.', '').replace(',', '').isdigit()):
                        score += 2
                
                # Strong indicators this is a DATA row (not header)
                # 1. Long numeric codes (barcodes)
                if cell.isdigit() and len(cell) > 8:
                    is_likely_data = True
                    score -= 10
                
                # 2. Currency values
                if 'ksh' in cell_lower or '$' in cell or '€' in cell or '£' in cell:
                    if any(c.isdigit() for c in cell):
                        is_likely_data = True
                        score -= 8
                
                # 3. Single digit numbers (likely IDs in data rows)
                if cell.isdigit() and len(cell) == 1:
                    score -= 3
                
                # 4. Product-like descriptions with specific details
                if '(' in cell and ')' in cell and len(cell) > 15:
                    # Like "Ameer Al Arab (Black) Edp" - likely data
                    score -= 2
            
            # Don't consider rows with clear data indicators as headers
            if not is_likely_data and score > 0:
                header_candidates.append((analysis['idx'], score, analysis))
        
        # Select best header candidate
        if header_candidates:
            header_candidates.sort(key=lambda x: x[1], reverse=True)
            best_score = header_candidates[0][1]
            
            if best_score >= 6:  # Increased threshold for confidence
                structure['header_row_idx'] = header_candidates[0][0]
                structure['data_start_idx'] = header_candidates[0][0] + 1
        
        # If still no header found, look for patterns
        if structure['header_row_idx'] is None:
            # Find first row after titles that has text but minimal barcodes/numbers
            for analysis in row_analysis:
                if analysis['type'] in ['title', 'empty']:
                    continue
                
                # Headers typically have more text than pure numbers/barcodes
                if (analysis['text_count'] >= 3 and 
                    analysis['barcode_count'] == 0 and
                    analysis['short_number_count'] <= 1):
                    structure['header_row_idx'] = analysis['idx']
                    structure['data_start_idx'] = analysis['idx'] + 1
                    break
        
        # Last resort: use first non-title row
        if structure['header_row_idx'] is None:
            for analysis in row_analysis:
                if analysis['type'] not in ['title', 'empty']:
                    structure['header_row_idx'] = analysis['idx']
                    structure['data_start_idx'] = analysis['idx'] + 1
                    break
        
        # Determine column count
        structure['column_count'] = max(len(row) for row in table) if table else 0
        
        # Check if data has sequential IDs in first column
        if structure['data_start_idx'] and structure['data_start_idx'] < len(table):
            first_col_values = []
            for idx in range(structure['data_start_idx'], min(structure['data_start_idx'] + 5, len(table))):
                if idx < len(table) and len(table[idx]) > 0:
                    val = str(table[idx][0]).strip()
                    if val.isdigit():
                        first_col_values.append(int(val))
            
            # Check if sequential
            if len(first_col_values) >= 3:
                if first_col_values == list(range(first_col_values[0], first_col_values[0] + len(first_col_values))):
                    structure['has_sequential_ids'] = True
        
        return structure
    
    def create_headers(row, col_count, structure=None):
        """Create clean, unique headers from a row with intelligent naming."""
        headers = []
        seen = {}
        
        for col_idx in range(col_count):
            header = clean_header(row[col_idx]) if col_idx < len(row) else None
            
            if not header:
                # Generate intelligent column names based on position
                if col_idx == 0 and structure and structure.get('has_sequential_ids'):
                    header = "id"
                else:
                    header = f"column_{col_idx + 1}"
            else:
                # Ensure uniqueness by adding suffix if needed
                original = header
                counter = 1
                while header in seen:
                    header = f"{original}_{counter}"
                    counter += 1
                seen[header] = True
            
            headers.append(header)
        return headers
    
    converted_files = []
    
    if merge:
        # Merge all tables into a single JSON file with table metadata
        output_path = os.path.join(output_dir, f"{base_filename}.json")
        
        # First, check if these are actually valid tables or just poorly parsed text
        # A valid table should have:
        # 1. Multiple columns (at least 2-3)
        # 2. Multiple rows
        # 3. Consistent structure
        is_valid_table_data = False
        for table in tables:
            if len(table) >= 3:  # At least 3 rows
                # Check column count consistency
                col_counts = [len(row) for row in table]
                max_cols = max(col_counts) if col_counts else 0
                
                # If most rows have multiple columns and structure is consistent
                if max_cols >= 3:
                    # Check if at least 70% of rows have 2+ columns
                    multi_col_rows = sum(1 for count in col_counts if count >= 2)
                    if multi_col_rows / len(table) >= 0.7:
                        is_valid_table_data = True
                        break
        
        # If tables look like poorly parsed text, fall back to text extraction
        if not is_valid_table_data:
            if pdf_path and os.path.exists(pdf_path):
                result = extract_structured_text_json(pdf_path)
                has_valid_tables = True  # Mark as valid so file gets written
            else:
                result = {"tables": []}
                has_valid_tables = False
            
            # Write the text extraction result and skip table processing
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            converted_files.append(output_path)
        else:
            # Strategy: Find the first table with valid headers, use those headers for ALL subsequent tables
            master_headers = None
            all_data_rows = []
            
            for table_idx, table in enumerate(tables, start=1):
                if len(table) < 1:
                    continue
                
                # Analyze table structure
                structure = analyze_table_structure(table)
                
                if not structure:
                    continue
                
                # If we don't have master headers yet, try to get them from this table
                if master_headers is None and structure['header_row_idx'] is not None:
                    header_row_idx = structure['header_row_idx']
                    master_headers = create_headers(table[header_row_idx], structure['column_count'], structure)
                    print(f"[INFO] Using headers from table {table_idx}: {master_headers}")
                    
                    # Get data from this table (if any)
                    data_start_idx = structure['data_start_idx']
                    if data_start_idx < len(table):
                        for row_idx in range(data_start_idx, len(table)):
                            row = table[row_idx]
                            if any(str(cell).strip() for cell in row if cell):
                                all_data_rows.append(row)
                else:
                    # We already have master headers, just collect data rows
                    # Start from the beginning if no headers found in this table
                    if structure['header_row_idx'] is None:
                        start_idx = 0
                        print(f"[INFO] Table {table_idx} has no headers, treating all rows as data")
                    else:
                        # Skip the header row in this table since we're using master headers
                        start_idx = structure['data_start_idx'] if structure['data_start_idx'] < len(table) else 0
                        print(f"[INFO] Table {table_idx} has headers, but using master headers, starting from row {start_idx}")
                    
                    for row_idx in range(start_idx, len(table)):
                        row = table[row_idx]
                        # Skip empty rows
                        if not any(str(cell).strip() for cell in row if cell):
                            continue
                        # Skip rows that look like headers (contain header keywords)
                        row_str = ' '.join([str(cell).lower() for cell in row if cell])
                        if any(keyword in row_str for keyword in ['sno', 'barcode', 'product', 'image', 'brand', 'description', 'wholesale', 'retail']):
                            # Check if it's actually a header row (not data with these words in product names)
                            is_header = False
                            for cell in row[:3]:  # Check first 3 cells
                                cell_str = str(cell).lower().strip()
                                if cell_str in ['sno', 's.no', 'no', 'barcode', 'bar code', 'product image', 'image']:
                                    is_header = True
                                    break
                            if is_header:
                                print(f"[INFO] Skipping duplicate header row in table {table_idx}")
                                continue
                        
                        all_data_rows.append(row)
            
            # If we found headers and data, create the merged result
            if master_headers and all_data_rows:
                has_valid_tables = True
                
                # Convert all data rows to dictionaries using master headers
                table_data = []
                for row in all_data_rows:
                    row_dict = {}
                    for col_idx, header in enumerate(master_headers):
                        value = row[col_idx] if col_idx < len(row) else ""
                        value = str(value).strip() if value else ""
                        row_dict[header] = value
                    
                    table_data.append(row_dict)
                
                result = {
                    "tables": [{
                        "table_number": 1,
                        "rows": len(table_data),
                        "columns": len(master_headers),
                        "headers": master_headers,
                        "data": table_data
                    }]
                }
            else:
                result = {"tables": []}
            
            # If no valid tables found in table mode, extract as structured text
            if not has_valid_tables or not result["tables"]:
                if pdf_path and os.path.exists(pdf_path):
                    result = extract_structured_text_json(pdf_path)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            converted_files.append(output_path)
    else:
        # Save each table as a separate JSON file
        has_valid_tables = False
        
        for idx, table in enumerate(tables, start=1):
            if len(table) < 2:
                continue
            
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.json")
            
            # Analyze table structure intelligently
            structure = analyze_table_structure(table)
            
            if not structure or structure['header_row_idx'] is None:
                continue
            
            has_valid_tables = True
            header_row_idx = structure['header_row_idx']
            data_start_idx = structure['data_start_idx']
            
            if data_start_idx >= len(table):
                continue
            
            # Extract title if present
            title_text = None
            if structure['title_rows']:
                title_parts = []
                for title_analysis in structure['title_rows']:
                    non_empty = [cell for cell in title_analysis['cells'] if cell]
                    title_parts.extend(non_empty)
                title_text = ' '.join(title_parts) if title_parts else None
            
            # Create headers
            headers = create_headers(table[header_row_idx], structure['column_count'], structure)
            
            # Convert data rows to dictionaries
            table_data = []
            for row_idx in range(data_start_idx, len(table)):
                row = table[row_idx]
                
                # Skip empty rows
                if not any(str(cell).strip() for cell in row if cell):
                    continue
                
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    value = row[col_idx] if col_idx < len(row) else ""
                    value = str(value).strip() if value else ""
                    row_dict[header] = value
                
                table_data.append(row_dict)
            
            if table_data:
                result = {
                    "table_number": idx,
                    "rows": len(table_data),
                    "columns": len(headers),
                    "headers": headers,
                    "data": table_data
                }
                
                if title_text:
                    result["title"] = title_text
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(result, f, indent=2, ensure_ascii=False)
                
                converted_files.append(output_path)
        
        # If no valid tables found, create a single text file
        if not has_valid_tables and pdf_path and os.path.exists(pdf_path):
            output_path = os.path.join(output_dir, f"{base_filename}.json")
            result = extract_structured_text_json(pdf_path)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            converted_files.append(output_path)
    
    return converted_files
    """
    Save extracted tables to JSON files.
    Returns list of created file paths.
    """
    
    def clean_header(header):
        """Clean and normalize header text."""
        if not header or not str(header).strip():
            return None
        # Remove newlines and extra whitespace
        cleaned = str(header).replace('\n', ' ').strip()
        # Remove multiple spaces
        cleaned = ' '.join(cleaned.split())
        return cleaned if cleaned else None
    
    def is_title_row(row):
        """
        Detect if a row is likely a title/header row rather than column headers.
        Title rows typically have:
        1. Only one or two cells with content (rest are empty)
        2. Longer text that spans conceptually across the table
        3. No data-like patterns (numbers, codes, structured text)
        """
        non_empty_cells = [cell for cell in row if cell and str(cell).strip()]
        
        # If most cells are empty, likely a title
        if len(non_empty_cells) <= 2 and len(row) > 3:
            # Check if the non-empty content is longer text (likely a title)
            for cell in non_empty_cells:
                cell_text = str(cell).strip()
                # Titles are usually longer and contain spaces or newlines
                if len(cell_text) > 15 or '\n' in str(cell):
                    return True
        
        return False
    
    def is_header_row(row):
        """
        Detect if a row contains column headers vs actual data.
        Headers typically:
        1. Contain descriptive text (not just numbers or codes)
        2. Have words like: Sno, Name, Description, Price, Code, Item, etc.
        3. Are not primarily numeric or barcode-like
        """
        if not row:
            return False
        
        non_empty_cells = [str(cell).strip() for cell in row if cell and str(cell).strip()]
        
        if len(non_empty_cells) < 3:
            return False
        
        # Common header keywords
        header_keywords = [
            'sno', 'serial', 'number', 'name', 'description', 'price', 'code', 
            'barcode', 'brand', 'item', 'product', 'quantity', 'qty', 'amount',
            'date', 'time', 'category', 'type', 'status', 'id', 'image',
            'wholesale', 'retail', 'ml', 'pcs', 'ctn', 'carton', 'pieces'
        ]
        
        # Check if cells contain header-like text
        header_score = 0
        numeric_score = 0
        
        for cell in non_empty_cells[:5]:  # Check first 5 cells
            cell_lower = str(cell).lower().strip()
            
            # Check for header keywords
            if any(keyword in cell_lower for keyword in header_keywords):
                header_score += 2
            
            # Check if it's mostly text (not numbers or codes)
            if len(cell_lower) > 0:
                # If it contains letters and isn't just a barcode-like number
                if any(c.isalpha() for c in cell_lower):
                    if not cell_lower.replace('\n', '').replace(' ', '').isdigit():
                        header_score += 1
                
                # Penalize if it's purely numeric or looks like data
                if cell_lower.isdigit() or (len(cell_lower) > 8 and cell_lower.isdigit()):
                    numeric_score += 1
        
        # Row is a header if it has more header characteristics than data characteristics
        return header_score > numeric_score and header_score >= 2
    
    def find_header_row(table):
        """
        Find the actual header row in a table, skipping title rows.
        Returns (header_row_index, title_text or None)
        """
        title_text = None
        title_rows = []
        
        for idx, row in enumerate(table):
            # Collect title rows at the beginning
            if is_title_row(row):
                non_empty = [str(cell).strip() for cell in row if cell and str(cell).strip()]
                if non_empty:
                    title_rows.extend(non_empty)
                continue
            
            # Check if this row is a header row
            if is_header_row(row):
                title_text = ' '.join(title_rows) if title_rows else None
                return idx, title_text
        
        # If no header found, assume first non-title row is header
        first_non_title = 0
        for idx, row in enumerate(table):
            if not is_title_row(row):
                first_non_title = idx
                break
        
        title_text = ' '.join(title_rows) if title_rows else None
        return first_non_title, title_text
    
    def create_headers(row, col_count):
        """Create clean, unique headers from a row."""
        headers = []
        seen = {}
        for col_idx in range(col_count):
            header = clean_header(row[col_idx]) if col_idx < len(row) else None
            
            if not header:
                # Use column position for empty headers
                header = f"column_{col_idx + 1}"
            else:
                # Ensure uniqueness by adding suffix if needed
                original = header
                counter = 1
                while header in seen:
                    header = f"{original}_{counter}"
                    counter += 1
                seen[header] = True
            
            headers.append(header)
        return headers
    
    converted_files = []
    
    if merge:
        # Merge all tables into a single JSON file with table metadata
        output_path = os.path.join(output_dir, f"{base_filename}.json")
        
        # Convert tables to structured JSON with table separation
        result = {
            "tables": []
        }
        
        for table_idx, table in enumerate(tables, start=1):
            if len(table) < 2:  # Skip tables with only header or empty tables
                continue
            
            # Find the actual header row, skipping titles
            header_row_idx, title = find_header_row(table)
            
            if header_row_idx >= len(table):
                continue  # No valid data rows
            
            # Determine column count from the table
            max_cols = max(len(row) for row in table) if table else 0
            
            # Create headers from the identified header row
            headers = create_headers(table[header_row_idx], max_cols)
            
            # Convert rows to dictionaries (skip title and header rows)
            table_data = []
            for row_idx in range(header_row_idx + 1, len(table)):
                row = table[row_idx]
                
                # Skip empty rows
                if not any(str(cell).strip() for cell in row if cell):
                    continue
                
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    value = row[col_idx] if col_idx < len(row) else ""
                    # Convert to string and clean
                    value = str(value).strip() if value else ""
                    row_dict[header] = value
                
                table_data.append(row_dict)
            
            if table_data:  # Only add non-empty tables
                table_info = {
                    "table_number": table_idx,
                    "rows": len(table_data),
                    "columns": len(headers),
                    "headers": headers,
                    "data": table_data
                }
                
                # Add title if found
                if title:
                    table_info["title"] = title
                
                result["tables"].append(table_info)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        
        converted_files.append(output_path)
    else:
        # Save each table as a separate JSON file
        for idx, table in enumerate(tables, start=1):
            if len(table) < 2:  # Skip tables with only header or empty tables
                continue
            
            output_path = os.path.join(output_dir, f"{base_filename}_table{idx}.json")
            
            # Find the actual header row, skipping titles
            header_row_idx, title = find_header_row(table)
            
            if header_row_idx >= len(table):
                continue  # No valid data rows
            
            # Determine column count
            max_cols = max(len(row) for row in table) if table else 0
            
            # Create headers from the identified header row
            headers = create_headers(table[header_row_idx], max_cols)
            
            # Convert rows to dictionaries (skip title and header rows)
            table_data = []
            for row_idx in range(header_row_idx + 1, len(table)):
                row = table[row_idx]
                
                # Skip empty rows
                if not any(str(cell).strip() for cell in row if cell):
                    continue
                
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    value = row[col_idx] if col_idx < len(row) else ""
                    # Convert to string and clean
                    value = str(value).strip() if value else ""
                    row_dict[header] = value
                
                table_data.append(row_dict)
            
            if table_data:  # Only save non-empty tables
                result = {
                    "table_number": idx,
                    "rows": len(table_data),
                    "columns": len(headers),
                    "headers": headers,
                    "data": table_data
                }
                
                # Add title if found
                if title:
                    result["title"] = title
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(result, f, indent=2, ensure_ascii=False)
                
                converted_files.append(output_path)
    
    return converted_files


def process_conversion(job_id, file_infos, parser, merge, output_format='csv'):
    """
    Background worker to process PDF conversion.
    Updates job status as it progresses.
    Supports output formats: 'csv' or 'excel'
    """
    job = conversion_jobs[job_id]
    job['status'] = 'processing'
    job['progress'] = 0
    
    try:
        total_files = len(file_infos)
        all_converted = []
        
        for idx, file_info in enumerate(file_infos):
            file_id = file_info['fileId']
            filename = file_info['filename']
            
            # Find the PDF file
            pdf_path = None
            for f in os.listdir(UPLOAD_FOLDER):
                if f.startswith(file_id):
                    pdf_path = os.path.join(UPLOAD_FOLDER, f)
                    break
            
            if not pdf_path or not os.path.exists(pdf_path):
                job['errors'].append(f"File not found: {filename}")
                continue
            
            # Update status
            job['status'] = 'converting'
            job['currentFile'] = filename
            
            # Extract tables
            if parser == 'pdfplumber':
                tables = extract_tables_pdfplumber(pdf_path)
                
                # Fallback to text if no tables found
                if not tables:
                    tables = extract_text_lines(pdf_path)
            else:
                # Future: Add tabula support
                tables = extract_tables_pdfplumber(pdf_path)
            
            # Create output directory for this file
            base_filename = os.path.splitext(filename)[0]
            file_output_dir = os.path.join(CONVERTED_FOLDER, job_id, file_id)
            os.makedirs(file_output_dir, exist_ok=True)
            
            # Save to output format (CSV, Excel, or JSON)
            # For JSON format, always call save function as it handles text extraction fallback
            if output_format == 'json':
                converted_files = save_tables_to_json(
                    tables, 
                    file_output_dir, 
                    base_filename, 
                    merge,
                    pdf_path
                )
                
                for csv_path in converted_files:
                    file_info = {
                        'fileId': f"{file_id}_{os.path.basename(csv_path)}",
                        'originalFileId': file_id,
                        'filename': os.path.basename(csv_path),
                        'filepath': csv_path,
                        'size': os.path.getsize(csv_path)
                    }
                    all_converted.append(file_info)
            elif tables:
                if output_format == 'excel':
                    converted_files = save_tables_to_excel(
                        tables, 
                        file_output_dir, 
                        base_filename, 
                        merge
                    )
                else:  # CSV
                    converted_files = save_tables_to_csv(
                        tables, 
                        file_output_dir, 
                        base_filename, 
                        merge
                    )
                
                for csv_path in converted_files:
                    file_info = {
                        'fileId': f"{file_id}_{os.path.basename(csv_path)}",
                        'originalFileId': file_id,
                        'filename': os.path.basename(csv_path),
                        'filepath': csv_path,
                        'size': os.path.getsize(csv_path)
                    }
                    all_converted.append(file_info)
            else:
                job['errors'].append(f"No tables extracted from: {filename}")
            
            # Update progress
            job['progress'] = int(((idx + 1) / total_files) * 100)
        
        # Mark as completed
        job['status'] = 'completed'
        job['progress'] = 100
        job['convertedFiles'] = all_converted
        job['completedAt'] = datetime.utcnow().isoformat()
        job['message'] = f"Successfully converted {len(all_converted)} file(s)"
        
    except Exception as e:
        job['status'] = 'error'
        job['progress'] = 100
        job['error'] = str(e)
        job['message'] = f"Conversion failed: {str(e)}"


@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'service': 'conversion',
        'timestamp': datetime.utcnow().isoformat()
    })


@app.route('/api/convert', methods=['POST'])
def convert():
    """
    Start PDF to CSV/Excel/JSON conversion job.
    
    Request body:
    {
        "fileIds": ["abc123", "def456"],
        "parser": "pdfplumber",  // or "tabula"
        "merge": false,
        "outputFormat": "csv"  // or "excel" or "json"
    }
    """
    data = request.get_json()
    
    if not data:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INVALID_REQUEST',
                'message': 'Request body is required'
            }
        }), 400
    
    file_ids = data.get('fileIds', [])
    parser = data.get('parser', 'pdfplumber')
    merge = data.get('merge', False)
    output_format = data.get('outputFormat', 'csv')
    
    if not file_ids:
        return jsonify({
            'success': False,
            'error': {
                'code': 'NO_FILES',
                'message': 'No files provided for conversion'
            }
        }), 400
    
    # Create conversion job
    job_id = uuid.uuid4().hex
    
    # Build file info list (assuming we get fileIds from upload service)
    file_infos = []
    for file_id in file_ids:
        # In production, we would fetch metadata from upload service
        # For now, we'll construct minimal info
        file_infos.append({
            'fileId': file_id,
            'filename': f"{file_id}.pdf"  # Will be updated when we find the actual file
        })
    
    conversion_jobs[job_id] = {
        'jobId': job_id,
        'status': 'pending',
        'progress': 0,
        'fileIds': file_ids,
        'parser': parser,
        'merge': merge,
        'outputFormat': output_format,
        'createdAt': datetime.utcnow().isoformat(),
        'currentFile': None,
        'convertedFiles': [],
        'errors': [],
        'error': None,
        'message': 'Conversion queued'
    }
    
    # Start background conversion
    thread = threading.Thread(
        target=process_conversion,
        args=(job_id, file_infos, parser, merge, output_format),
        daemon=True
    )
    thread.start()
    
    return jsonify({
        'success': True,
        'data': {
            'jobId': job_id,
            'status': 'pending',
            'message': 'Conversion started'
        },
        'timestamp': datetime.utcnow().isoformat()
    }), 200


@app.route('/api/status/<job_id>', methods=['GET'])
def get_status(job_id):
    """Get conversion job status."""
    if job_id not in conversion_jobs:
        return jsonify({
            'success': False,
            'error': {
                'code': 'JOB_NOT_FOUND',
                'message': 'Conversion job not found'
            }
        }), 404
    
    job = conversion_jobs[job_id]
    
    return jsonify({
        'success': True,
        'data': {
            'jobId': job['jobId'],
            'status': job['status'],
            'progress': job['progress'],
            'message': job.get('message', ''),
            'currentFile': job.get('currentFile'),
            'convertedFiles': job.get('convertedFiles', []),
            'errors': job.get('errors', []),
            'error': job.get('error'),
            'createdAt': job['createdAt'],
            'completedAt': job.get('completedAt')
        },
        'timestamp': datetime.utcnow().isoformat()
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002, debug=True)
